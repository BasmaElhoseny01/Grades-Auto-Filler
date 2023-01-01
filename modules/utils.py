# Imports
import cv2
import numpy as np
import pandas as pd
import skimage.io as io
import matplotlib.pyplot as plt
from skimage.color import rgb2gray, rgb2hsv
import math
from skimage.transform import resize
from openpyxl import load_workbook
import pytesseract
from sklearn import svm
from skimage.feature import hog

# from skimage.morphology import thin
from sklearn.datasets import fetch_openml
from sklearn.model_selection import train_test_split,GridSearchCV

import os
import imutils


def HogFun(img_1):
    resized_img_1 = resize(img_1, (128*4, 64*4))

    # creating hog features
    fd_1, hog_image_1 = hog(resized_img_1, orientations=9, pixels_per_cell=(
        8, 8), cells_per_block=(2, 2), visualize=True, multichannel=False)
    # show_images([img_1,resized_img_1,hog_image_1],["Gray","Resized","hog_image"])
    # Fd Faeture
    #     print(np.shape(fd_1))
    return fd_1, hog_image_1


def show_images(images, titles=None):
    # This function is used to show image(s) with titles by sending an array of images and an array of associated titles.
    # images[0] will be drawn with the title titles[0] if exists
    # You aren't required to understand this function, use it as-is.
    n_ims = len(images)
    if titles is None:
        titles = ['(%d)' % i for i in range(1, n_ims + 1)]
    fig = plt.figure()
    n = 1
    for image, title in zip(images, titles):
        a = fig.add_subplot(1, n_ims, n)
        if image.ndim == 2:
            plt.gray()
        plt.imshow(image)
        a.set_title(title)
        n += 1
    fig.set_size_inches(np.array(fig.get_size_inches()) * n_ims)
    plt.show()


def showHist(img):
    # An "interface" to matplotlib.axes.Axes.hist() method
    plt.figure()
    imgHist = histogram(img, nbins=256)

    bar(imgHist[1].astype(np.uint8), imgHist[0], width=0.8, align='center')


# Reorder 4 Points in clockwise order, starting from top left
def reorderPoints(points):

    points = points.reshape((4, 2))
    newPoints = np.zeros((4, 1, 2), dtype=np.int32)
    add = points.sum(1)

    newPoints[0] = points[np.argmin(add)]
    newPoints[3] = points[np.argmax(add)]
    diff = np.diff(points, axis=1)
    newPoints[1] = points[np.argmin(diff)]
    newPoints[2] = points[np.argmax(diff)]

    return newPoints
# Darw Reactangle


def drawRectangle(img, biggest, thickness):
    cv2.line(img, (biggest[0][0][0], biggest[0][0][1]),
             (biggest[1][0][0], biggest[1][0][1]), (0, 255, 0), thickness)
    cv2.line(img, (biggest[0][0][0], biggest[0][0][1]),
             (biggest[2][0][0], biggest[2][0][1]), (0, 255, 0), thickness)
    cv2.line(img, (biggest[3][0][0], biggest[3][0][1]),
             (biggest[2][0][0], biggest[2][0][1]), (0, 255, 0), thickness)
    cv2.line(img, (biggest[3][0][0], biggest[3][0][1]),
             (biggest[1][0][0], biggest[1][0][1]), (0, 255, 0), thickness)

    return img


def fill_grades(file_path, id, marks):
    """
    This function fills the excel sheet with the marks to the student with the given id.
    Arguments:
        file_path: String
        id: Integer
        marks: Array
    """

    # reading the file
    file = load_workbook(file_path)
    sheet = file.active

    # Extracting some info
    num_of_questions = len(marks)
    id_location = 0

    # searching for the id
    for index, cell in enumerate(sheet['A']):
        if (cell.value == id):
            id_location = index + 1
            break

    # if the id was not found, throw an exception
    if (id_location == 0):
        raise Exception("Id is not found")

    # assigning the marks to the row
    counter = 0
    for cell in (sheet.cell(row=id_location, column=i) for i in range(4, 4 + num_of_questions)):
        cell.value = marks[counter]
        counter = counter + 1

#     for index, cell in enumerate(sheet[id_location]):
#         if(index >= 3):
#             cell.value = marks[counter]
#             counter = counter + 1

    # save the file with a new name
    file.save(file_path)
    return True
