# Imports
import cv2 as cv
import numpy as np
import skimage.io as io
import matplotlib.pyplot as plt
import math
from openpyxl import load_workbook

# Read RGB image uisng CV and result is a RGB image


def Read_RGB_Image(path):
    # cv2.IMREAD_COLOR: It specifies to load a color image. Any transparency of image will be neglected.
    image = cv.imread(path, 1)
    # Converting BGR color to RGB color format
    RGB_img = cv.cvtColor(image, cv.COLOR_BGR2RGB)
    return RGB_img


def show_images(images, titles=None):
    # This function is used to show image(s) with titles by sending an array of images and an array of associated titles.
    # images[0] will be drawn with the title titles[0] if exists
    # You aren't required to understand this function, use it as-is.
    n_ims = len(images)
    if titles is None:
        titles = ['(%d)' % i for i in range(1, n_ims + 1)]
    fig = plt.figure()
    n = 1
    for image, title in zip(images, titles):
        a = fig.add_subplot(1, n_ims, n)
        if image.ndim == 2:
            plt.gray()
        plt.imshow(image)
        a.set_title(title)
        n += 1
    fig.set_size_inches(np.array(fig.get_size_inches()) * n_ims)
    plt.show()


def showHist(img):
    # An "interface" to matplotlib.axes.Axes.hist() method
    plt.figure()
    imgHist = histogram(img, nbins=256)

    bar(imgHist[1].astype(np.uint8), imgHist[0], width=0.8, align='center')

# # Show images in plot
# def Show_Images(images,titles):
#     # Create Figure
#     fig=plt.figure(figsize=(20, 10))

#     # Setting number of rows and columns
#     columns=2 # 2columns
#     rows=math.ceil(np.shape(images)[0]/columns) #No of rows

#     n=1 #Counter for the subplots order
#     for image,title in zip(images,titles):
#         #Add subplot @ order postion
#         fig.add_subplot(rows, columns, n)
#         n+=1
#         plt.axis('off')
#         plt.title(title)
#         plt.imshow(image)


# Function to sort points in a contour in clockwise order, starting from top left
def processContour(approx):
    # Reshape array([x, y], ...) to array( array([x], [y]), ...)
    approx = approx.reshape((4, 2))

    # Sort points in clockwise order, starting from top left
    pts = np.zeros((4, 2), dtype=np.float32)

    # Add up all values
    # Smallest sum = top left point
    # Largest sum = bottom right point
    s = approx.sum(axis=1)
    pts[0] = approx[np.argmin(s)]
    pts[2] = approx[np.argmax(s)]

    # For the other 2 points, compute difference between all points
    # Smallest difference = top right point
    # Largest difference = bottom left point
    diff = np.diff(approx, axis=1)
    pts[1] = approx[np.argmin(diff)]
    pts[3] = approx[np.argmax(diff)]

    # Calculate smallest height and width
    width = int(min(pts[1][0] - pts[0][0], pts[2][0] - pts[3][0]))
    height = int(min(pts[3][1] - pts[0][1], pts[2][1] - pts[1][1]))

    return pts, width, height

# Function to find the largest 4-sided contour from an array of countours


def findLargestQuadrilateralContour(contours):
    # Sort contours from smallest area to biggest
    sorted_contours = sorted(contours, key=cv2.contourArea, reverse=True)

    biggest_contour = None
    biggest_contour_approx = None

    for cnt in sorted_contours:
        # Get the length of the perimeter
        perimeter = cv2.arcLength(cnt, True)

        # Approximate a shape that resembles the contour
        # This is needed because the image might be warped, thus
        # edges are curved and not perfectly straight
        approx = cv2.approxPolyDP(cnt, 0.01 * perimeter, True)

        # Check if the approximation contains only 4 sides
        # (i.e. quadrilateral)
        if len(approx) == 4:
            biggest_contour = cnt
            biggest_contour_approx = approx
            break

    return [biggest_contour], [biggest_contour_approx]


def fill_grades(file_path, id, marks):
    """
    This function fills the excel sheet with the marks to the student with the given id.
    Arguments:
        file_path: String
        id: Integer
        marks: Array
    """

    # reading the file
    file = load_workbook(file_path)
    sheet = file.active

    # Extracting some info
    num_of_questions = len(marks)
    id_location = 0

    # searching for the id
    for index, cell in enumerate(sheet['A']):
        if (cell.value == id):
            id_location = index + 1
            break

    # if the id was not found, throw an exception
    if (id_location == 0):
        raise Exception("Id is not found")

    # assigning the marks to the row
    counter = 0
    for cell in (sheet.cell(row=id_location, column=i) for i in range(4, 4 + num_of_questions)):
        cell.value = marks[counter]
        counter = counter + 1

#     for index, cell in enumerate(sheet[id_location]):
#         if(index >= 3):
#             cell.value = marks[counter]
#             counter = counter + 1

    # save the file with a new name
    file.save(file_path)
    return True
